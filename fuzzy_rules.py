#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import openpyxl
from openpyxl import Workbook
from pathlib import Path
import numpy as np

xlsx_file = Path( 'norm_rtmatrix.xlsx')
xlsx_file2 = Path( 'norm_tpmatrix.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)
wb_obj2 = openpyxl.load_workbook(xlsx_file2)
# Read the active sheet:
sheet = wb_obj.active
sheet2 = wb_obj2.active
book = Workbook()
sheet1 = book.active
j=0
k=1
count=0
count2=0
for i in range(3,sheet.max_row+1):
	#print(i)
	for j in range(1,sheet.max_column+1):
		c1=sheet.cell(row=i,column=j)
		sheet1.cell(row=k,column=1).value = c1.value
		c2=sheet2.cell(row=i,column=j)
		sheet1.cell(row=k,column=2).value = c2.value
		if ((c2.value > 0.04) or (c1.value > 0 and c1.value <= 0.02)): 
			ts=1
			count+=1
		elif ((c2.value > 0.03 and c2.value <=0.04) or (c1.value > 0.02 and c1.value <= 0.040)):
			ts=1
			count+=1
		elif ((c2.value > 0.02 and c2.value <=0.03) and (c1.value > 0.040 and c1.value <=0.06)):
			ts=0
			count2+=1
		elif((c2.value > 0.01 and c2.value <=0.02) and (c1.value > 0.06 and c1.value <= 0.08)):
			ts=0
			count2+=1
		elif (c2.value <= 0.01 and c1.value >0.08):
			ts=0
			count2+=1
		elif ((c2.value > 0.02 and c2.value <=0.03) and (c1.value > 0.06 and c1.value <=0.08)):
			ts=0
			count2+=1
		sheet1.cell(row=k,column=3).value = ts
		k=k+1
print("Pre-processed file saved as text.xlsx")
book.save('text.xlsx')

