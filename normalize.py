#!/usr/bin/env python
# coding: utf-8

# In[17]:


import openpyxl
from openpyxl import Workbook
from pathlib import Path
import numpy as np

filename='rtmatrix.xlsx'
xlsx_file = Path(filename)
wb_obj = openpyxl.load_workbook(xlsx_file) 

# Read the active sheet:
sheet = wb_obj.active
j=0
maxi=0.0
min1=np.zeros(sheet.max_row)
max1=np.zeros(sheet.max_row)
mini=0.0
cell=0
for i in range(1,sheet.max_row+1):
	if(sheet.cell(row=i,column=1).value != None):
		mini=sheet.cell(row=i,column=1).value
	maxi=0.0
	for j in range(1,sheet.max_column+1):
		if(sheet.cell(row=i,column=j).value != None):
			cell = float(sheet.cell(row=i,column=j).value)
		if(maxi<cell):
			maxi=cell
		if(mini>cell and cell != -1):
			mini=cell
	min1[i-1]=mini
	max1[i-1]=maxi
	#print(i)
book=Workbook()
sheet1=book.active
#sheet.max_row+1
for k in range(1,sheet.max_row+1):
	for v in range(1,sheet.max_column+1):
		if(sheet.cell(row=k,column=v).value != None):
			#print(sheet.cell(row=k,column=v),min1[k-1],max1[k-1])
			sheet1.cell(row=k,column=v).value=(float(sheet.cell(row=k,column=v).value)-min1[k-1])/(max1[k-1]-min1[k-1])
book.save('norm_rtmatrix.xlsx')

print(filename, " normalized and saved")

# In[ ]:




